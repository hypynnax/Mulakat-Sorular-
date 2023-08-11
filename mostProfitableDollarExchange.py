import requests
import sqlite3
import time
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

#create customer class
class Customer:
    def __init__(self, name, balance=0):
        self.name = name
        self.balance = balance

#class that holds all operations to be performed
class Process:
    #excel operations
    #excel creation
    def create_excel(title, *headers):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title

        area = 'A1'
        for header in headers:
            sheet[area] = header
            area = chr(ord(area[0])+1)+'1'

        workbook.save('C:\\Users\\hypyn\\Desktop\\dolar karı\\'+title+'.xlsx')
    
    #save data to excel
    def excel_print():
        workbook = load_workbook("Dolar Kuru.xlsx")
        sheet = workbook.active

        current_date = datetime.today()
        row = 2

        turkish_months = {
            1: "ocak",
            2: "subat",
            3: "mart",
            4: "nisan",
            5: "mayis",
            6: "haziran",
            7: "temmuz",
            8: "agustos",
            9: "eylul",
            10: "ekim",
            11: "kasim",
            12: "aralik"
        }

        while row <= 32:
            formatted_date = f"{current_date.day}-{turkish_months[current_date.month]}-{current_date.year}"
            try:
                url = f"https://bigpara.hurriyet.com.tr/doviz/merkez-bankasi-doviz-kurlari/{formatted_date}/"

                response = requests.get(url)
                if response.status_code == 200:
                    try:
                        soup = BeautifulSoup(response.content, 'html.parser')
                        table = soup.find("div", {"class": "tableBox"})
                        rows = table.find_all("ul")
                        columns = rows[1].find_all("li")
                        formatted_date = current_date.strftime('%d.%m.%Y')
                        usd_rate_buying = columns[2].text
                        usd_rate_sales = columns[3].text
                        sheet.cell(row=row, column=1, value=formatted_date)
                        sheet.cell(row=row, column=2, value=usd_rate_buying)
                        sheet.cell(row=row, column=3, value=usd_rate_sales)
                        row += 1
                    except Exception as ex:
                        pass
                else:
                    print('Web page could not be reached. Status code:', response.status_code)

                current_date -= timedelta(days=1)
            except Exception as ex:
                print("Connection Error",ex)
        
        workbook.save('Dolar Kuru.xlsx')
    
    #reading data from excel
    def excel_read():
        workbook = load_workbook("Dolar Kuru.xlsx")
        sheet = workbook.active

        date_list = []
        buyin_rate_list = []
        sales_rate_list = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date_list.append(row[0])
            buyin_rate_list.append(row[1])
            sales_rate_list.append(row[2])

        return date_list, buyin_rate_list, sales_rate_list
    
    #checking excel
    def excel_exist():
        try:
            workbook = load_workbook("Dolar Kuru.xlsx")
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                if row[0]:
                    return True
        except Exception as ex:
            pass
        return False

    #database operations
    #database creation
    def create_database(database_name, table_name):
        conn = sqlite3.connect(database_name+'.db')
        cursor = conn.cursor()
        cmd = "CREATE TABLE IF NOT EXISTS " + table_name + " (name, balance)"
        cursor.execute(cmd)
        conn.commit()
    
    #save to database
    def database_registration(database_name, table_name, name, balance):
        conn = sqlite3.connect(database_name+'.db')
        cursor = conn.cursor()
        cursor.execute("INSERT INTO " + table_name + " (name, balance) VALUES (?, ?)", (name, balance))
        conn.commit()
        conn.close()    
    
    #pull data from database
    def pull_data(database_name, table_name, query):
        conn = sqlite3.connect(database_name+'.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM " + table_name + " WHERE name = " + query)
        result = cursor.fetchall()[0][1]
        conn.close()
        return result

start_time = time.time()
Process.create_database("User Information", "Users")
while True:
    answer = input("ARE YOU A REGISTERED USER? Y/N ").lower().strip()
    customer = Customer(input("Enter your name ").capitalize())
    if answer == 'y' or answer == "yes":
        customer.balance = Process.pull_data("User Information", "Users", "name")
        break
    elif answer == 'n' or answer == "no":
        customer.balance = int(input("Enter the amount you want to deposit (TL) "))
        Process.database_registration("User Information", "Users", customer.name, customer.balance)
        break
    else:
        print("You entered an invalid value")

answer = input("Do you want to update dollar rates? Yes/No ").lower().strip()
if not Process.excel_exist() or answer == "yes" or answer == "y" or answer == "evet" or answer == "e" or answer == "1":
    Process.create_excel('Dolar Kuru', 'Tarih', 'Dolar Kuru(Alis)', 'Dolar Kuru(Satis)')
    Process.excel_print()
dates, buyinges, saleses = Process.excel_read()
received_dollars = 0.0
rate = 0
max_profit = 0
max_profit_buying_date = datetime.today()
max_profit_sales_date = datetime.today()
for i in range(30):
    for j in range(30):
        buying = float(buyinges[i].replace(',', '.'))
        sales = float(saleses[j].replace(',', '.'))
        if customer.balance/buying*sales > max_profit:
            max_profit = customer.balance/buying*sales
            received_dollars = customer.balance/buying
            max_profit_buying_date = dates[i]
            max_profit_sales_date = dates[j]
rate = (max_profit - customer.balance)*100/customer.balance

print("If you had, the dollar exchange you would buy on {0} and sell on {1} would bring you the highest profit. "
    "With your TL balance of {2:.2f}, you would buy {3:.2f} $ and the profit rate would be %{4:.2f}."
    .format(max_profit_buying_date, max_profit_sales_date, customer.balance, received_dollars, rate))

finish_time = time.time()
run_time = finish_time - start_time
print("Program run time: {:.2f} second".format(run_time))
